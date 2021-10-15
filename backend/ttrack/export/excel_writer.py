# -*- coding: utf-8 -*-
"""Export the data of the data store into Excel reports while preserving manually entered data"""
import logging
import os

import shutil

import datetime

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Border, Side, colors, Font
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter

from utils.errors import ExcelWriterError

logger = logging.getLogger(__name__)

class ExcelWriter:
    REPORT_MILAGE = 'milage'
    REPORT_INCOME = 'income'

    """Handle the data export to MS Excel."""
    def __init__(self, config):
        """The initializer of the class.
        
        :type config: backend.ttrack.utils.config_reader.ConfigReader
        :param config: the current configuration of ttrack 
        """
        self._milage_filename = config.export_milage
        self._income_filename = config.export_income
        self._export_path = config.export_path

    def backup_and_create(self, data_store, report_type):
        """Create a backup of the current reports and re-export the data.
        
        :type data_store: backend.ttrack.persistence.data_store.DataStore
        :param data_store: the current data to export
        :return: None
        """
        try:
            self._backup_files()
            self._create_export(data_store, report_type)
        except IOError as io_err:
            logger.info('IO error received: {0}'.format(io_err.strerror))
            msg = 'Error during IO operation: {0}'.format(io_err.strerror)
            raise ExcelWriterError(msg)
        except ExcelWriterError:
            raise ExcelWriterError
        except Exception as err:
            logger.info('unexpected error received: {0}'.format(err.message))
            msg = 'Reveived unexpected error: {0}'.format(err.message)
            raise ExcelWriterError(msg)

    def _backup_files(self):
        """Create a backup of the current reports."""
        try:
            backup_prefix = datetime.datetime.now().strftime("%Y%m%d_%H%M%S_")
            if not os.path.exists(self._export_path):
                os.makedirs(self._export_path)
            if os.path.exists(self._milage_filename):
                backup_name = os.path.join(self._export_path, backup_prefix + os.path.basename(self._milage_filename))
                shutil.copy(self._milage_filename, backup_name)
            if os.path.exists(self._income_filename):
                backup_name = os.path.join(self._export_path, backup_prefix + os.path.basename(self._income_filename))
                shutil.copy(self._income_filename, backup_name)
        except (OSError, IOError) as err:
            logger.info('os error received: {0}, {1}'.format(err.strerror, err.filename))
            msg = 'Unable to backup file: {0}, {1}'.format(err.strerror, err.filename)
            raise ExcelWriterError(msg)

    def _create_export(self, data_store, report_type):
        """Export the current data of data_store into MS Excel reports.

               :type data_store: backend.ttrack.persistence.data_store.DataStore
               :param data_store: the data to be exported
               :return: None
       """
        if report_type == self.REPORT_MILAGE:
            self._create_milage_export(data_store)
        elif report_type == self.REPORT_INCOME:
            self._create_income_export(data_store)
        else:
            logger.info('wrong report type used: {0}'.format(report_type))

    def _create_milage_export(self, data_store):
        """Create the milage report as MS Excel file.
        
        :type data_store: backend.ttrack.persistence.data_store.DataStore
        :param data_store: the data to be exported
        :return: None
        """
        try:
            wb = load_workbook(self._milage_filename)
            wb = self._generate_milage_rows(wb, data_store)
            wb.save(self._milage_filename)
        except Exception as err:
            raise ExcelWriterError(err.message)

    def _create_income_export(self, data_store):
        """Create the income/expense report as MS Excel file.
        
        :type data_store: backend.ttrack.persistence.data_store.DataStore
        :param data_store: the data to be exported
        :return: None
        """
        try:
            wb = load_workbook(self._income_filename)
            wb = self._generate_expense_rows(wb, data_store)
            wb.save(self._income_filename)
        except Exception as err:
            raise ExcelWriterError(err.message)

    def _find_start_marker(self, max_row, sheet):
        """Find the row with the 'start_generated' marker. 
        
        :type max_row: int
        :param max_row: the maximum row number in the worksheet
        :type sheet: openpyxl.worksheet.worksheet.Worksheet
        :param sheet: the worksheet to be processed
        :return: number of row with start marker
        """
        start_row = 0
        for row_num in range(1, max_row):
            cell_name = 'A{0}'.format(row_num)
            if sheet[cell_name].value == 'start_generated':
                start_row = row_num
                break
        if start_row == 0:
            logger.info('start_generated marker not found')
            msg = 'Invalid worksheet: no start_generated marker found'
            raise ExcelWriterError(msg)
        return start_row

    def _create_milage_header(self, month_name, sheet):
        sheet['E2'].value = month_name + ' ' + '2019'

    def _generate_milage_rows(self, wb, data_store):
        template_ws = wb.get_sheet_by_name('Vorlage')
        # check if for every month a sheet is avaiable
        for month_num, month_name in [(1, u'Jaenner'), (2, u'Februar'), (3, u'Maerz'), (4, u'April'), (5, u'Mai'),
                                      (6, u'Juni'), (7, u'Juli'), (8, u'August'), (9, u'September'),
                                      (10, u'Oktober'), (11, u'November'), (12, u'Dezember')]:

            data = data_store.get_milage_data(month_num, '2019')

            # if sheet is not found, just copy the template and set the date
            if not month_name in wb.sheetnames:
                template_sheet = wb.copy_worksheet(template_ws)
                self._create_milage_header(month_name, template_sheet)
                template_sheet.title = month_name

            # copy all manually entered values and remove the rest
            old_sheet = wb.get_sheet_by_name(month_name)
            max_row = old_sheet.max_row
            max_col = old_sheet.max_column

            new_sheet = wb.copy_worksheet(template_ws)
            new_sheet.page_setup.orientation = new_sheet.ORIENTATION_LANDSCAPE
            self._create_milage_header(month_name, new_sheet)

            old_start_row = self._find_start_marker(max_row, old_sheet)
            for row_num in range(old_start_row + 1, max_row+1):
                cell_name = 'A{0}'.format(row_num)
                if old_sheet[cell_name].value == 'generated':
                    # skip this row
                    pass
                else:
                    # save the values
                    manual_value = ()
                    for col_num in range(1, max_col + 1):
                        manual_value = manual_value + (old_sheet.cell(row=row_num, column=col_num).value,)
                    if not manual_value[1] is None:
                        data.append(manual_value)

            data = sorted(data, key=lambda value: value[1])

            new_start_row = self._find_start_marker(max_row, new_sheet)
            next_row = new_start_row + 1
            for row_num in range(0, len(data)):
                # copy this row
                for col_num in range(1, max_col + 1):
                    new_sheet.cell(row=next_row, column=col_num).value = data[row_num][col_num-1]
                    if col_num == 2:
                        new_sheet.cell(row=next_row, column=col_num).number_format = 'dd.mm.yy'
                    if col_num != 1:
                        new_sheet.cell(row=next_row, column=col_num).border = Border(left=Side(border_style='thin',
                                                                                           color='FF000000'),
                                                                                 right=Side(border_style='thin',
                                                                                            color='FF000000'),
                                                                                 top=Side(border_style='thin',
                                                                                          color='FF000000'),
                                                                                 bottom=Side(border_style='thin',
                                                                                             color='FF000000'))
                next_row = next_row + 1
            # correct the sum forumlar
            new_sheet["E3"] = "=SUM(F{0}:F{1})".format(new_start_row + 1, new_sheet.max_row)
            # rename the sheet
            wb.remove(old_sheet)
            new_sheet.title = month_name
            new_sheet.column_dimensions['A'].hidden = True

        return wb

    def _generate_expense_rows(self, wb, data_store):
        # get list of income and expense entries for the current year, sorted by date
        # loop over all sheets and then for each sheet
        # look at B of start_row, if there is a date -> this is manually entered data
        # compare the date with the first date of the list and add it on the appropriate place

        template_ws = wb.get_sheet_by_name('Vorlage')
        # check if for every month a sheet is avaiable
        win_loss_row_idx = 0
        for month_num, month_name in [(1, u'Jaenner'), (2, u'Februar'), (3, u'Maerz'), (4, u'April'), (5, u'Mai'),
                                      (6, u'Juni'), (7, u'Juli'), (8, u'August'), (9, u'September'),
                                      (10, u'Oktober'), (11, u'November'), (12, u'Dezember')]:

            data = data_store.get_expense_data(month_num, '2019')

            # if sheet is not found, just copy the template and set the date
            if not month_name in wb.sheetnames:
                template_sheet = wb.copy_worksheet(template_ws)
                self._create_income_expense_header(month_name, template_sheet)
                template_sheet.title = month_name

            # copy all manually entered values and remove the rest
            old_sheet = wb.get_sheet_by_name(month_name)
            max_row = old_sheet.max_row
            max_col = template_ws.max_column

            new_sheet = wb.copy_worksheet(template_ws)
            # set landscape orientation
            new_sheet.page_setup.orientation = new_sheet.ORIENTATION_LANDSCAPE
            self._create_income_expense_header(month_name, new_sheet)

            old_start_row = self._find_start_marker(max_row, old_sheet)
            for row_num in range(old_start_row + 1, max_row + 1):
                cell_name = 'A{0}'.format(row_num)
                if old_sheet[cell_name].value == 'generated':
                    # skip this row
                    pass
                elif old_sheet[cell_name].value in ['sum_income', 'sum_expense']:
                    # stop parsing
                    break
                else:
                    # save the values
                    manual_value = ()
                    for col_num in range(1, max_col + 1):
                        manual_value = manual_value + (old_sheet.cell(row=row_num, column=col_num).value,)
                    if not manual_value[1] is None:
                        data.append(manual_value)

            data = sorted(data, key=lambda value: value[1])

            new_start_row = self._find_start_marker(max_row, new_sheet)
            next_row = new_start_row + 1
            for row_num in range(0, len(data)):
                # copy this row
                for col_num in range(1, max_col + 1):
                    new_sheet.cell(row=next_row, column=col_num).value = data[row_num][col_num - 1]
                    if col_num == 2:
                        new_sheet.cell(row=next_row, column=col_num).number_format = 'dd.mm.yy'
                    if col_num != 1:
                        new_sheet.cell(row=next_row, column=col_num).border = Border(left=Side(border_style='thin',
                                                                                               color='FF000000'),
                                                                                     right=Side(border_style='thin',
                                                                                                color='FF000000'))
                next_row = next_row + 1

            if (next_row > new_start_row + 1): # the loop was executed at least once -> reduce the index
                next_row -= 1

            # add footer
            self._create_income_expense_footer(new_sheet, month_name, new_start_row + 1, next_row, win_loss_row_idx)
            win_loss_row_idx = next_row + 3

            # rename the sheet
            wb.remove(old_sheet)
            new_sheet.title = month_name
            new_sheet.column_dimensions['A'].hidden = True

        return wb

    def _create_income_expense_header(self, month_name, sheet):
        sheet['D1'].value = month_name + ' ' + '2019'
        self._style_range(sheet, 'E3:E3', Border(top=Side(border_style='thin', color=colors.BLACK),
                                                     left=Side(border_style='thin', color=colors.BLACK),
                                                     bottom=Side(border_style='thin', color=colors.BLACK),
                                                     right=Side(border_style='thin', color=colors.BLACK)))
        self._style_range(sheet, 'F3:L3', Border(top=Side(border_style='thin', color=colors.BLACK),
                                                 left=Side(border_style='thin', color=colors.BLACK),
                                                 bottom=Side(border_style='thin', color=colors.BLACK),
                                                 right=Side(border_style='thin', color=colors.BLACK)))


    def _create_income_expense_footer(self, sheet, month_name, first_row, last_row, last_month_win_loss_row):

        # create border after last row
        self._style_range(sheet, 'B{0}:L{0}'.format(last_row + 1), Border(top=Side(border_style='thin',
                                                                                   color=colors.BLACK)))
        # create sum income row
        sheet['A{0}'.format(last_row + 1)].value = 'sum_income'
        range_string = 'B{0}:C{0}'.format(last_row + 1)
        sheet.merge_cells(range_string)
        sheet[range_string.split(":")[0]].value = 'Summe Einnahmen Monat'
        sheet['D{0}'.format(last_row + 1)] = '=SUM(E{0}:E{1})'.format(last_row + 1, last_row + 1)
        sheet['E{0}'.format(last_row + 1)] = '=SUM(E{0}:E{1})'.format(first_row, last_row)
        self._style_range(sheet, 'B{0}:E{0}'.format(last_row + 1), Border(top=Side(border_style='thin',
                                                                                   color=colors.BLACK),
                                                 left=Side(border_style='thin', color=colors.BLACK),
                                                 bottom=Side(border_style='thin', color=colors.BLACK),
                                                 right=Side(border_style='thin', color=colors.BLACK)))

        # create sum expense row
        sheet['A{0}'.format(last_row + 2)].value = 'sum_expense'
        range_string = 'B{0}:C{0}'.format(last_row + 2)
        sheet.merge_cells(range_string)
        sheet[range_string.split(":")[0]].value = 'Summe Ausgaben Monat'
        sheet['D{0}'.format(last_row + 2)] = '=SUM(F{0}:L{0})*-1'.format(last_row + 2)
        sheet['F{0}'.format(last_row + 2)] = '=SUM(F{0}:F{1})'.format(first_row, last_row)
        sheet['G{0}'.format(last_row + 2)] = '=SUM(G{0}:G{1})'.format(first_row, last_row)
        sheet['H{0}'.format(last_row + 2)] = '=SUM(H{0}:H{1})'.format(first_row, last_row)
        sheet['I{0}'.format(last_row + 2)] = '=SUM(I{0}:I{1})'.format(first_row, last_row)
        sheet['J{0}'.format(last_row + 2)] = '=SUM(J{0}:J{1})'.format(first_row, last_row)
        sheet['K{0}'.format(last_row + 2)] = '=SUM(K{0}:K{1})'.format(first_row, last_row)
        sheet['L{0}'.format(last_row + 2)] = '=SUM(L{0}:L{1})'.format(first_row, last_row)
        self._style_range(sheet, 'B{0}:D{0}'.format(last_row + 2), Border(top=Side(border_style='thin',
                                                                                   color=colors.BLACK),
                                                                          left=Side(border_style='thin',
                                                                                    color=colors.BLACK),
                                                                          bottom=Side(border_style='thin',
                                                                                      color=colors.BLACK),
                                                                          right=Side(border_style='thin',
                                                                                     color=colors.BLACK)))
        self._style_range(sheet, 'F{0}:L{0}'.format(last_row + 2), Border(top=Side(border_style='thin',
                                                                                   color=colors.BLACK),
                                                                          left=Side(border_style='thin',
                                                                                    color=colors.BLACK),
                                                                          bottom=Side(border_style='thin',
                                                                                      color=colors.BLACK),
                                                                          right=Side(border_style='thin',
                                                                                     color=colors.BLACK)))

        # create win / loss row
        range_string = 'B{0}:C{0}'.format(last_row + 3)
        sheet.merge_cells(range_string)
        cell = sheet[range_string.split(":")[0]]
        cell.value = 'Gewinn/Verlust'
        cell.font = Font(bold=True)
        sheet['D{0}'.format(last_row + 3)] = '=SUM(D{0}:D{1})'.format(last_row + 1, last_row + 2)
        sheet['D{0}'.format(last_row + 3)].border = Border(top=Side(border_style='thin',
                                                                                   color=colors.BLACK),
                                                                          left=Side(border_style='medium',
                                                                                    color=colors.BLACK),
                                                                          bottom=Side(border_style='medium',
                                                                                      color=colors.BLACK),
                                                                          right=Side(border_style='thin',
                                                                                     color=colors.BLACK))
        sheet['D{0}'.format(last_row + 3)].font = Font(bold=True)
        sheet.conditional_formatting.add('D{0}'.format(last_row + 3),
                                         CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True,
                                                    font=Font(bold=True, color=colors.RED)))

        # set number format for the entire range
        for idx in range(first_row, last_row + 3 + 1, 1):
            for col_idx in range(4, 16 + 1):
                col = get_column_letter(col_idx)
                sheet['{0}{1}'.format(col, idx)].number_format = '#,##0.00_-'

        # add sheet wide sum (starting at month 2)
        if month_name !=  u'Jaenner':
            last_month_name = u''
            if month_name == u'Februar':
                last_month_name = u'Jaenner'
            elif month_name == u'Maerz':
                last_month_name = u'Februar'
            elif month_name == u'April':
                last_month_name = u'Maerz'
            elif month_name == u'Mai':
                last_month_name = u'April'
            elif month_name == u'Juni':
                last_month_name = u'Mai'
            elif month_name == u'Juli':
                last_month_name = u'Juni'
            elif month_name == u'August':
                last_month_name = u'Juli'
            elif month_name == u'September':
                last_month_name = u'August'
            elif month_name == u'Oktober':
                last_month_name = u'September'
            elif month_name == u'November':
                last_month_name = u'Oktober'
            elif month_name == u'Dezember':
                last_month_name = u'November'

            ic_row = last_row + 5
            ec_row = ic_row + 1
            wlc_row = ec_row + 1

            if month_name == u'Februar':
                ic_row_last = last_month_win_loss_row - 2
                ec_row_last = ic_row_last + 1
                wlc_row_last = ec_row_last + 1
            else:
                ic_row_last = last_month_win_loss_row + 2
                ec_row_last = ic_row_last + 1
                wlc_row_last = ec_row_last + 1


            # add sheet wide income sum row
            range_string = 'B{0}:C{0}'.format(ic_row)
            sheet.merge_cells(range_string)
            sheet[range_string.split(":")[0]].value = 'Summe Einnahmen kumuliert'
            sheet['D{0}'.format(ic_row)] = '=SUM(E{0}:E{0})'.format(ic_row)
            formula = '=E{0} + {1}!E{2}'.format(last_row + 1, last_month_name.encode('utf-8'),
                                                                       ic_row_last)
            sheet['E{0}'.format(ic_row)] = formula
            self._style_range(sheet, 'D{0}:E{0}'.format(ic_row), Border(top=Side(border_style='thin',
                                                                                 color=colors.BLACK),
                                                                        left=Side(border_style='thin',
                                                                                  color=colors.BLACK),
                                                                        bottom=Side(border_style='thin',
                                                                                    color=colors.BLACK),
                                                                        right=Side(border_style='thin',
                                                                                   color=colors.BLACK)))
            # add sheet wide expense sum row
            range_string = 'B{0}:C{0}'.format(ec_row)
            sheet.merge_cells(range_string)
            sheet[range_string.split(":")[0]].value = 'Summe Ausgaben kumuliert'
            sheet['D{0}'.format(ec_row)] = '=SUM(F{0}:L{0})*-1'.format(ec_row)
            formula = '=F{0} + {1}!F{2}'.format(last_row + 2, last_month_name.encode('utf-8'), ec_row_last)
            sheet['F{0}'.format(ec_row)] = formula
            formula = '=G{0} + {1}!G{2}'.format(last_row + 2, last_month_name.encode('utf-8'), ec_row_last)
            sheet['G{0}'.format(ec_row)] = formula
            formula = '=H{0} + {1}!H{2}'.format(last_row + 2, last_month_name.encode('utf-8'), ec_row_last)
            sheet['H{0}'.format(ec_row)] = formula
            formula = '=I{0} + {1}!I{2}'.format(last_row + 2, last_month_name.encode('utf-8'), ec_row_last)
            sheet['I{0}'.format(ec_row)] = formula
            formula = '=J{0} + {1}!J{2}'.format(last_row + 2, last_month_name.encode('utf-8'), ec_row_last)
            sheet['J{0}'.format(ec_row)] = formula
            formula = '=K{0} + {1}!K{2}'.format(last_row + 2, last_month_name.encode('utf-8'), ec_row_last)
            sheet['K{0}'.format(ec_row)] = formula
            formula = '=L{0} + {1}!L{2}'.format(last_row + 2, last_month_name.encode('utf-8'), ec_row_last)
            sheet['L{0}'.format(ec_row)] = formula
            self._style_range(sheet, 'B{0}:D{0}'.format(ec_row), Border(top=Side(border_style='thin',
                                                                                 color=colors.BLACK),
                                                                        left=Side(border_style='thin',
                                                                                  color=colors.BLACK),
                                                                        bottom=Side(border_style='thin',
                                                                                    color=colors.BLACK),
                                                                        right=Side(border_style='thin',
                                                                                   color=colors.BLACK)))
            self._style_range(sheet, 'F{0}:L{0}'.format(ec_row), Border(top=Side(border_style='thin',
                                                                                 color=colors.BLACK),
                                                                        left=Side(border_style='thin',
                                                                                  color=colors.BLACK),
                                                                        bottom=Side(border_style='thin',
                                                                                    color=colors.BLACK),
                                                                        right=Side(border_style='thin',
                                                                                   color=colors.BLACK)))

            # add sheet wide win loss row
            range_string = 'B{0}:C{0}'.format(wlc_row)
            sheet.merge_cells(range_string)
            sheet[range_string.split(":")[0]].value = 'Gewinn/Verlust kumuliert'
            sheet[range_string.split(":")[0]].font = Font(bold=True)
            sheet['D{0}'.format(wlc_row)] = '=SUM(D{0}:D{1})'.format(ic_row, ec_row)

            sheet['D{0}'.format(wlc_row)].font = Font(bold=True)
            sheet['D{0}'.format(wlc_row)].border = Border(top=Side(border_style='thin', color=colors.BLACK),
                                                          left=Side(border_style='medium', color=colors.BLACK),
                                                          bottom=Side(border_style='medium', color=colors.BLACK),
                                                          right=Side(border_style='thin', color=colors.BLACK))
            sheet.conditional_formatting.add('D{0}'.format(wlc_row),
                                             CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True,
                                                        font=Font(bold=True, color=colors.RED)))

            # set number format for the entire range
            for idx in range(ic_row, wlc_row + 1, 1):
                for col_idx in range(4, 16 + 1):
                    col = get_column_letter(col_idx)
                    sheet['{0}{1}'.format(col, idx)].number_format = '#,##0.00_-'


    def _style_range(self, ws, cell_range, border=None):
        """
        :param ws:  Excel worksheet instance
        :param range: An excel range to style (e.g. A1:F20)
        :param border: An openpyxl Border object
        """

        start_cell, end_cell = cell_range.split(':')
        start_coord = coordinate_from_string(start_cell)
        start_row = start_coord[1]
        start_col = column_index_from_string(start_coord[0])
        end_coord = coordinate_from_string(end_cell)
        end_row = end_coord[1]
        end_col = column_index_from_string(end_coord[0])

        for row in range(start_row, end_row + 1):
            for col_idx in range(start_col, end_col + 1):
                col = get_column_letter(col_idx)
                #ws.cell('%s%s' % (col, row)).border = border
                ws['%s%s' % (col, row)].border = border

