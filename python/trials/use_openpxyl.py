# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet import Worksheet

wb = load_workbook('Fahrtenbuch.xlsx')
print wb.get_sheet_names()
ws = wb.get_sheet_by_name(u'Jänner')
print ws['C6'].value
print ws['C7'].value
print ws['C8'].value
print ws['C9'].value
myCell = Cell()
myCell.number_format = 'dd.mm.yy'